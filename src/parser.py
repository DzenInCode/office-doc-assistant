"""Parse xlsx/csv/pdf documents into plain text for LLM consumption."""
from __future__ import annotations

import io
import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv", ".pdf"}


def parse(filename: str, data: bytes) -> str:
    """Parse a document into a plain-text representation.

    Args:
        filename: Original filename (used to detect type by extension).
        data: Raw file bytes.

    Returns:
        Plain-text representation of the document.

    Raises:
        ValueError: If the file extension is not supported.
    """
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type {ext!r}. "
            f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    if ext in (".xlsx", ".xls"):
        return _parse_xlsx(data)
    if ext == ".csv":
        return _parse_csv(data)
    if ext == ".pdf":
        return _parse_pdf(data)
    raise ValueError(f"Unhandled extension {ext!r}")  # defensive


def _parse_xlsx(data: bytes) -> str:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append("\t".join(cells).rstrip())
        while rows and not rows[-1].strip():
            rows.pop()
        parts.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))
    return "\n\n".join(parts)


def _parse_csv(data: bytes) -> str:
    df = pd.read_csv(io.BytesIO(data))
    return df.to_csv(index=False)


def _parse_pdf(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    pages: list[str] = []
    for i, page in enumerate(reader.pages, 1):
        text = page.extract_text() or ""
        pages.append(f"## Page {i}\n{text}")
    return "\n\n".join(pages)
